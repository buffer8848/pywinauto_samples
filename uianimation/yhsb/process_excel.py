#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# auth: libinfei/179770346@qq.com
# time: 2018/12/09
# desc: 处理excel表

#--------------------------------------------------------------------------------------------------

def Merge_excels(heduijingzhi_excel, toucun_excel, zichan_excel, jijin_excel, outputexcel, email_server_url,
    email_server_port, sender_email, sender_passwd, reciever_email, exPath, flagCir):
    print(1)
    import xlrd
    import openpyxl
    from common import send_email_to_admin
    #先打开核对净值总表
    hedui_data = openpyxl.load_workbook(heduijingzhi_excel)
    print(1.1)

    #得到头寸excel表
    hedui_toucun_sheet = hedui_data.create_sheet("估值-头寸预测表") #追加的方式
    toucun_data = xlrd.open_workbook(toucun_excel)
    toucun_sheet = toucun_data.sheet_by_index(0)
    for row in range(1, toucun_sheet.nrows):
        for col in range(1, toucun_sheet.ncols):
            hedui_toucun_sheet.cell(row, col, toucun_sheet.cell_value(row, col))
    print(2)

    #合并估值-资产情况表
    hedui_zichan_sheet = hedui_data.create_sheet("估值-资产情况表") #追加的方式
    zichan_data = xlrd.open_workbook(zichan_excel)
    zichan_sheet = zichan_data.sheet_by_index(0)
    for row in range(1, zichan_sheet.nrows):
        for col in range(1, zichan_sheet.ncols):
            hedui_zichan_sheet.cell(row, col, zichan_sheet.cell_value(row, col))
    print(3)

    # 恒生-基金资产表
    hedui_jijin_sheet = hedui_data.create_sheet("恒生-基金资产表")  # 追加的方式
    jijin_data = xlrd.open_workbook(jijin_excel)
    jijin_sheet = jijin_data.sheet_by_index(0)
    for row in range(1, jijin_sheet.nrows):
        for col in range(1, jijin_sheet.ncols):
            hedui_jijin_sheet.cell(row, col, jijin_sheet.cell_value(row, col))
    print(4)
    try:
        hedui_data.save(outputexcel)
    except Exception as e:
        hedui_data.save(outputexcel)
        print(e)
    send_email_to_admin("helloworld", email_server_url, email_server_port, sender_email, sender_passwd,
                        reciever_email, exPath + "\\净值核对结果.xlsx", 2, flagCir)
    print(5)


if __name__ == "__main__":
    heduijingzhi_excel = r"D:\work\uiauto\yhsb\uianimation\核对净值示范空表.xlsx"
    toucun_excel = r"D:\work\uiauto\yhsb\uianimation\180003银华-道琼斯88指数资产情况统计表.xls"
    zichan_excel = r"D:\work\uiauto\yhsb\uianimation\1800032018年12月05日银华-道琼斯88指数现金头寸预测汇总表.xls"
    outputexcel = r"D:\work\uiauto\yhsb\uianimation\核对净值表.xlsx"
    Merge_excels(heduijingzhi_excel, toucun_excel, zichan_excel, outputexcel)